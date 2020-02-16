import openpyxl
import  mysql.connector
import sys
import time
def file_written(query,templatepath,sheetname):
    try:
        template = templatepath
        workbk = openpyxl.load_workbook(template)
        try:
            cnx = mysql.connector.connect(user='root', password='root', host='localhost', database='epc')
#            print('connected to localhost')
            cursor = cnx.cursor()
            query = query
            cursor.execute(query)
            worksht = workbk[sheetname]
            rowv=0
            #for row in worksht.iter_rows(min_row=1, max_col=3, max_row=2):
            list_s = list(cursor)
            for s in list_s:
                columnv = 0
                rowv=rowv+1
                print(len(s))
                for i in range(0,len(s)):
                    columnv = columnv + 1
                    print(rowv + 1)
                    print(columnv)
                    print(s[i])
                    worksht.cell(row=rowv + 1, column=columnv).value = s[i]
                    #time.sleep(2)
        except mysql.connector.Error as ex:
            print(ex)
        finally:
            cnx.close()
            #sys.exit(0)
        workbk.save(templatepath)
        workbk.close()
    except NameError as ex:
        print(ex)

query_ntpq="""SELECT b.location ,b.dra_node ,b.server_role, a.draname ,b.ip,a.delay
FROM (select draname,delay from tbl_dra_ntpq ) a left join tbl_dra2 b
on  a.draname =b.hostname
order by draname;"""
file_written(query_ntpq,'C:\\abhishek\\project\\python\\DRA\\dra_template.xlsx',"NTPQ")

query_hastat = """
select b.location ,b.dra_node ,b.server_role,b.ip,a.draname, a.service, a.statu as Status
from tbl_dra_ha_mystat a left join  tbl_dra2 b
on  a.draname =b.hostname
order by draname ;
"""
file_written(query_hastat,'C:\\abhishek\\project\\python\\DRA\\dra_template.xlsx',"HA_Stat")

query_ra="""
SELECT
b.location ,b.dra_node ,b.server_role,b.ip,
a.draname, a.severity,IF(STRCMP(a.severity,"*C") = 0, "Critical", "Major") as Sev1, a.alarmname FROM tbl_dra_stat a left join
tbl_dra2 b
on  a.draname =b.hostname
order by draname ;
"""
file_written(query_ra,'C:\\abhishek\\project\\python\\DRA\\dra_template.xlsx',"Recent_Alarms")

query_syscheck ="""
SELECT
b.location ,b.dra_node ,b.server_role,b.ip,a.draname, a.class,a.statu as status
FROM tbl_dra_sys_check a left join tbl_dra2 b
on  a.draname =b.hostname order by draname ;
"""

file_written(query_syscheck,'C:\\abhishek\\project\\python\\DRA\\dra_template.xlsx',"Sys_Check")


query_disckspace ="""
select
b.location ,b.dra_node ,b.server_role,b.ip,
a.draname, a.mountname as Mount , a.disusage from tbl_dra_diskcheck a
left join tbl_dra2 b
on  a.draname =b.hostname order by draname ;
"""

file_written(query_disckspace,'C:\\abhishek\\project\\python\\DRA\\dra_template.xlsx',"Disk_Usage")