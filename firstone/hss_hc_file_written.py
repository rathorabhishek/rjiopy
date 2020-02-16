import openpyxl
import mysql.connector
import sys
#cnx = mysql.connector.connect(host='localhost', user='rjiladmin', password='TestServer@123', port='3307')

dat = sys.argv[1]
def file_written(query,templatepath,sheetname):
    try:
        template = templatepath
        workbk = openpyxl.load_workbook(template)
        try:
            cnx = mysql.connector.connect(host='localhost', user='rjiladmin', password='TestServer@123', port='3307', database='epc')
#            print('connected to localhost')
            cursor = cnx.cursor()
            query = query
            cursor.execute(query)
            worksht = workbk[sheetname]
            rowv=0
            #for row in worksht.iter_rows(min_row=1, max_col=3, max_row=2):
            list_s = list(cursor)
            for s in list_s:
                columnv = 0
                rowv=rowv+1
                print(len(s))
                for i in range(0,len(s)):
                    columnv = columnv + 1
                    print(rowv + 1)
                    print(columnv)
                    print(s[i])
                    worksht.cell(row=rowv + 1, column=columnv).value = s[i]
                    #time.sleep(2)
        except mysql.connector.Error as ex:
            print(ex)
        finally:
            cnx.close()
            #sys.exit(0)
        workbk.save(templatepath)
        workbk.close()
    except NameError as ex:
        print(ex)

query_disk_utill="""select gwname, filesystem, blocks, used_blck, avail_blck, used_blck_util, mountname
from epc.tbl_hss_disk_util where date_format(dt,'%Y-%m-%d') = '2019-10-20' ;"""
file_written(query_disk_utill,'C:\\mylog\\HC_Logs\\hss_template.xlsx',"Disk_Util")

query_free_mem = """ select gwname, remote, refid , delay, offse1t, jitter
from epc.tbl_hss_clock_synch  where date_format(dt,'%Y-%m-%d') = '2019-10-20'; """
file_written(query_free_mem,'C:\\mylog\\HC_Logs\\hss_template.xlsx','Free_Memory')


query_ntpq =""" select gwname, remote, refid , delay, offse1t, jitter
from epc.tbl_hss_clock_synch  where date_format(dt,'%Y-%m-%d') = '2019-10-20';"""
file_written(query_ntpq,'C:\\mylog\\HC_Logs\\hss_template.xlsx','NTPQ_P')


quer_sar =""" select gwname, time1, per_idle from epc.tbl_hss_procss_usage where date_format(dt,'%Y-%m-%d') = '"""+dat+"""'; """
file_written(quer_sar,'C:\\mylog\\HC_Logs\\hss_template.xlsx','SAR')


query_sys_memory = """ SELECT gwname, total, free, used, buffer FROM epc.tbl_hss_mem_status where date_format(dt,'%Y-%m-%d') = '2019-10-20';  """
file_written(query_sys_memory,'C:\\mylog\\HC_Logs\\hss_template.xlsx','System_Mem')


query_ntst3868 = """ select gwname, typen, gwip, gwip2, status1 from epc.ntstat_3868 where date_format(dt,'%Y-%m-%d') = '2019-10-20'; """
file_written(query_ntst3868,'C:\\mylog\\HC_Logs\\hss_template.xlsx','NtStat3868')


query_ntst30300 = """ select gwname, typen, gwip, gwip2, status1 from epc.ntstat_30300 where date_format(dt,'%Y-%m-%d') = '2019-10-20'; """
file_written(query_ntst30300,'C:\\mylog\\HC_Logs\\hss_template.xlsx','NtStat30300')



query_ntst16611 = """ select gwname, typen, gwip, gwip2, status1 from epc.ntstat_16611 where date_format(dt,'%Y-%m-%d') = '2019-10-20'; """
file_written(query_ntst16611,'C:\\mylog\\HC_Logs\\hss_template.xlsx','NtStat16611')


























