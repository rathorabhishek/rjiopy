import openpyxl
import mysql.connector
import sys


#dat = sys.argv[1]
def file_written(query,templatepath,sheetname):
    try:
        template = templatepath
        workbk = openpyxl.load_workbook(template)
        try:
            cnx = mysql.connector.connect(host='localhost', user='rjiladmin', password='TestServer@123', port='3307', database='epc')
#            print('connected to localhost')
            cursor = cnx.cursor()
            query = query
            cursor.execute(query)
            worksht = workbk[sheetname]
            rowv=0
            #for row in worksht.iter_rows(min_row=1, max_col=3, max_row=2):
            list_s = list(cursor)
            for s in list_s:
                columnv = 0
                rowv=rowv+1
                print(len(s))
                for i in range(0,len(s)):
                    columnv = columnv + 1
                    print(rowv + 1)
                    print(columnv)
                    print(s[i])
                    worksht.cell(row=rowv + 1, column=columnv).value = s[i]
                    #time.sleep(2)
        except mysql.connector.Error as ex:
            print(ex)
        finally:
            cnx.close()
            #sys.exit(0)
        workbk.save(templatepath)
        workbk.close()
    except NameError as ex:
        print(ex)

query_free_m="""select name,total, free, 100*(cast(Free as signed)/cast(Total as signed)) as freeRatio   from epc.lcs_free_m """
file_written(query_free_m,'C:\\mylog\\lcs_hc\\LCS_HC.xlsx',"Free_Memory")


query_free_m="""select Name, Filesystem, Mounted_on,Used_Per from epc.lcs_diskspace  ;"""
file_written(query_free_m,'C:\\mylog\\lcs_hc\\LCS_HC.xlsx',"DiskSpace")

query_free_m="""select Name, replace(replace(Replace(status1,'Status : [1;31m',''),'(local)',''),'status : [1;31m','') Stat from epc.lcs_systemcheck """
file_written(query_free_m,'C:\\mylog\\lcs_hc\\LCS_HC.xlsx',"SystemStatus")