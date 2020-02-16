import openpyxl
import  mysql.connector
import sys
import time


def file_written(query,templatepath,sheetname):
    try:
        template = templatepath
        workbk = openpyxl.load_workbook(template)
        try:
            cnx = mysql.connector.connect(host='localhost', user='rjiladmin', password='TestServer@123', port='3307')
            #cnx = mysql.connector.connect(user='root', password='root', host='localhost', database='epc')
#            print('connected to localhost')
            cursor = cnx.cursor()
            query = query
            cursor.execute(query)
            worksht = workbk[sheetname]
            rowv=0
            #for row in worksht.iter_rows(min_row=1, max_col=3, max_row=2):
            list_s = list(cursor)
            for s in list_s:
                columnv = 0
                rowv=rowv+1
                print(len(s))
                for i in range(0,len(s)):
                    columnv = columnv + 1
                    print(rowv + 1)
                    print(columnv)
                    print(s[i])
                    worksht.cell(row=rowv + 1, column=columnv).value = s[i]
                    #time.sleep(2)
        except mysql.connector.Error as ex:
            print(ex)
        finally:
            cnx.close()
            #sys.exit(0)
        workbk.save(templatepath)
        workbk.close()
    except NameError as ex:
        print(ex)

tempelatepath = 'C:/mylog/gw_healthcheck/Gw_Template.xlsx'

query_load=" SELECT nodename, node, CPU0_load, MEM0_LOAD, CPU1_load, MEM1_LOAD FROM epc.tbl_gw_load_sts ;"
file_written(query_load,tempelatepath,"Load_Status")

query_node_sts = "SELECT Nodename, node, SHELF, SLOT, ACT_STS, ADM_STS, OPR_STS, SVC_STS FROM epc.tbl_gw_node_sts ;"
file_written(query_node_sts,tempelatepath,"Node_Status")

query_pckge_info  = "SELECT nodename, dt, tm, size, filname FROM epc.tbl_gw_pckg_info ;"
file_written(query_pckge_info,tempelatepath,"Pckge_Info")


query_port_status = "SELECT nodename, node, port, typ, act_stats, adm_status, opr_status, descr FROM epc.tbl_gw_port_sts ;"
file_written(query_port_status,tempelatepath,"Port_Status")

